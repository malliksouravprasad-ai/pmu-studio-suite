"""
Excel Integration Service — read, write, and style Excel files using PMU branding.

All PMU apps that produce Excel outputs should use these utilities to ensure
consistent formatting across the tool suite.

PMU Color Palette:
  Primary dark blue  : 1F3864  (headers, key labels)
  Secondary blue     : 2E75B6  (sub-headers, accents)
  Section grey-blue  : D6DCE4  (section separators)
  Highlight yellow   : FFF2CC  (flagged / attention rows)
  Alert red          : C00000  (errors, critical values)
  Success green      : 70AD47  (compliant / on-track)
"""
from __future__ import annotations
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── PMU Brand Colors ──────────────────────────────────────────────────────────

PMU_COLORS = {
    "primary":    "1F3864",
    "secondary":  "2E75B6",
    "section":    "D6DCE4",
    "highlight":  "FFF2CC",
    "alert":      "C00000",
    "success":    "70AD47",
    "white":      "FFFFFF",
    "light_grey": "F2F2F2",
}

# Pre-built style objects
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(name="Calibri", bold=True, color="1F3864", size=10)
_BODY_FONT    = Font(name="Calibri", size=10)
_BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
_TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1F3864")

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
_SECTION_FILL = PatternFill("solid", fgColor="D6DCE4")
_HIGHLIGHT    = PatternFill("solid", fgColor="FFF2CC")
_ALERT_FILL   = PatternFill("solid", fgColor="FFE0E0")
_SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")

_WRAP_TOP  = Alignment(wrap_text=True, vertical="top")
_CENTER    = Alignment(horizontal="center", vertical="center")
_LEFT_MID  = Alignment(horizontal="left", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ── Workbook creation ─────────────────────────────────────────────────────────

def create_workbook(sheet_title: str = "Sheet1") -> Workbook:
    """Return a new Workbook with one sheet."""
    wb = Workbook()
    wb.active.title = sheet_title
    return wb


def add_sheet(wb: Workbook, title: str):
    """Add a sheet to an existing workbook."""
    return wb.create_sheet(title=title)


# ── Styling ───────────────────────────────────────────────────────────────────

def style_title_row(ws, title: str, row: int = 1, col_span: int = 1):
    """Write and style a report title row."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = _LEFT_MID
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 24


def style_header_row(
    ws,
    headers: list[str],
    row: int = 1,
    col_widths: list[int] = None,
    color: str = "1F3864",
):
    """Write and style a header row with the PMU primary color."""
    fill = PatternFill("solid", fgColor=color)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _CENTER
        if col_widths and col_idx <= len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[row].height = 22


def style_section_row(ws, label: str, row: int, col_count: int):
    """Write and style a section separator row."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = _LEFT_MID
    if col_count > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    ws.row_dimensions[row].height = 18


def style_data_row(ws, row: int, col_count: int, alternate: bool = False, status: str = None):
    """Apply body styling to a data row. Optionally alternate row shading."""
    fill = _ALT_FILL if alternate else None
    if status == "alert":
        fill = _ALERT_FILL
    elif status == "success":
        fill = _SUCCESS_FILL
    elif status == "highlight":
        fill = _HIGHLIGHT

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        if fill:
            cell.fill = fill
        cell.alignment = _WRAP_TOP


def set_cell(ws, row: int, col: int, value: Any, bold: bool = False,
             wrap: bool = True, center: bool = False, color: str = None):
    """Write a value to a cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _BOLD_FONT if bold else _BODY_FONT
    cell.alignment = _CENTER if center else (_WRAP_TOP if wrap else _LEFT_MID)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)
    return cell


# ── DataFrame → Sheet ─────────────────────────────────────────────────────────

def write_dataframe(
    ws,
    df,
    start_row: int = 2,
    include_headers: bool = False,
    alternate_rows: bool = True,
):
    """
    Write a pandas DataFrame to a worksheet starting at start_row.
    Headers are NOT written unless include_headers=True (use style_header_row separately).
    """
    if include_headers:
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = _BOLD_FONT
        start_row += 1

    for row_idx, (_, row) in enumerate(df.iterrows()):
        ws_row = start_row + row_idx
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=ws_row, column=col_idx, value=value).font = _BODY_FONT
        if alternate_rows:
            style_data_row(ws, ws_row, len(df.columns), alternate=(row_idx % 2 == 1))


# ── Column / row helpers ──────────────────────────────────────────────────────

def auto_width(ws, min_width: int = 8, max_width: int = 55):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=min_width
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def freeze_header(ws, freeze_at: str = "A2"):
    """Freeze panes so the header row stays visible while scrolling."""
    ws.freeze_panes = freeze_at


def set_col_width(ws, col: int, width: int):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Read ──────────────────────────────────────────────────────────────────────

def read_excel(path: str | Path, sheet_name: str | int = 0) -> "pd.DataFrame":
    """Read an Excel file into a pandas DataFrame."""
    import pandas as pd
    return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")


def load_wb(path: str | Path) -> Workbook:
    """Load an existing Excel workbook."""
    return load_workbook(str(path))


# ── Save ──────────────────────────────────────────────────────────────────────

def save_workbook(wb: Workbook, app_folder: str, filename: str) -> str:
    """
    Save workbook to PMU_Tools/outputs/<app_folder>/<filename>.
    Returns the absolute output path.
    """
    out_dir = Path(__file__).parent.parent / "outputs" / app_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    wb.save(str(out_path))
    return str(out_path)


def save_to_path(wb: Workbook, path: str | Path) -> str:
    """Save workbook to an explicit path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)
