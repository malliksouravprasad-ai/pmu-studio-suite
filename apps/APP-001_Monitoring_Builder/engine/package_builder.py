"""
Package Builder — assemble all monitoring framework components into outputs.

Generates:
  1. Excel data template (empty collection sheet + data dictionary + instructions)
  2. Validation config JSON (importable into APP-002 Data Processing Studio)
  3. KPI config JSON (importable into APP-003 Analytics Studio)
  4. Form structure JSON (for Google Forms creation or reference)
  5. Google Form (if credentials available)
"""
import io
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .monitoring_model import MonitoringJob, FieldDef, FIELD_TYPES


# ── Styles ────────────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=14, bold=True, color="1F3864")
_LABEL_FONT = Font(bold=True, size=10, color="1F3864")
_INFO_FONT  = Font(italic=True, size=9, color="595959")
_REQ_FILL   = PatternFill("solid", fgColor="FCE4D6")  # light orange = required
_OPT_FILL   = PatternFill("solid", fgColor="EBF3FB")  # light blue = optional
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _set_col_widths(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def _buf(wb: Workbook) -> io.BytesIO:
    b = io.BytesIO(); wb.save(b); b.seek(0); return b


# ── Excel Data Template ────────────────────────────────────────────────────────

def build_excel_template(job: MonitoringJob) -> io.BytesIO:
    """
    Build an empty Excel data collection template.

    Sheets:
      - Data Collection  — headers from schema, dropdowns for choice fields
      - Data Dictionary  — field-by-field documentation
      - Instructions     — how to fill out the form
    """
    wb   = Workbook()
    active_fields = [f for f in job.fields if f.enabled and not f.is_calculated]

    # ── Sheet 1: Data Collection ──────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data Collection"

    # Title row
    ws_data.merge_cells(f"A1:{get_column_letter(len(active_fields))}1")
    ws_data["A1"].value = f"{job.monitoring_name} — Data Collection Template"
    ws_data["A1"].font  = _TITLE_FONT
    ws_data.row_dimensions[1].height = 30

    # Info row
    from datetime import date
    ws_data["A2"].value = f"Project: {job.project_code}  |  Entity: {job.entity_type}  |  Generated: {date.today().strftime('%d %b %Y')}"
    ws_data["A2"].font  = _INFO_FONT
    ws_data.row_dimensions[2].height = 18

    # Blank row
    ws_data.row_dimensions[3].height = 6

    # Header row (row 4)
    for i, fld in enumerate(active_fields, start=1):
        col = get_column_letter(i)
        cell = ws_data.cell(row=4, column=i, value=fld.label or fld.name)
        cell.fill      = _REQ_FILL if fld.required else _OPT_FILL
        cell.font      = _LABEL_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER
        ws_data.column_dimensions[col].width = min(30, max(12, len(fld.label or fld.name) + 4))
        ws_data.row_dimensions[4].height = 32

    # Add dropdown validation for choice fields (rows 5–1000)
    for i, fld in enumerate(active_fields, start=1):
        if fld.data_type in ("choice", "boolean") and fld.choices:
            choices = fld.choices
            if fld.data_type == "boolean" and not choices:
                choices = ["Yes", "No"]
            if choices:
                formula = '"' + ",".join(choices) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=not fld.required)
                dv.sqref = f"{get_column_letter(i)}5:{get_column_letter(i)}1000"
                ws_data.add_data_validation(dv)

    ws_data.freeze_panes = "A5"

    # ── Sheet 2: Data Dictionary ──────────────────────────────────────────────
    ws_dict = wb.create_sheet("Data Dictionary")
    ws_dict["A1"].value = f"Data Dictionary — {job.monitoring_name}"
    ws_dict["A1"].font  = _TITLE_FONT

    dict_headers = ["#", "Column Name", "Label", "Type", "Required", "Choices / Format",
                    "Description", "Example", "Identifier"]
    for ci, hdr in enumerate(dict_headers, start=1):
        c = ws_dict.cell(row=3, column=ci, value=hdr)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_dict.row_dimensions[3].height = 24

    for ri, fld in enumerate(active_fields, start=4):
        ws_dict.cell(row=ri, column=1, value=ri - 3)
        ws_dict.cell(row=ri, column=2, value=fld.name)
        ws_dict.cell(row=ri, column=3, value=fld.label)
        ws_dict.cell(row=ri, column=4, value=FIELD_TYPES.get(fld.data_type, fld.data_type))
        ws_dict.cell(row=ri, column=5, value="Yes" if fld.required else "No")
        choices_str = " | ".join(fld.choices) if fld.choices else ""
        ws_dict.cell(row=ri, column=6, value=choices_str)
        ws_dict.cell(row=ri, column=7, value=fld.description)
        ws_dict.cell(row=ri, column=8, value=fld.example)
        ws_dict.cell(row=ri, column=9, value="Yes" if fld.is_identifier else "")
        if ri % 2 == 0:
            for ci in range(1, 10):
                ws_dict.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="F0F4FA")

    ws_dict.freeze_panes = "A4"
    _set_col_widths(ws_dict, {"A": 5, "B": 25, "C": 30, "D": 20, "E": 10, "F": 40, "G": 50, "H": 25, "I": 12})

    # ── Sheet 3: Instructions ─────────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instructions")
    ws_inst["A1"].value = "Instructions for Data Entry"
    ws_inst["A1"].font  = _TITLE_FONT
    ws_inst.column_dimensions["A"].width = 80

    instructions = [
        "",
        f"This template is for collecting {job.entity_type}-level data for: {job.monitoring_name}",
        "",
        "HOW TO FILL:",
        "• Do not change column headers in row 4 of the 'Data Collection' sheet.",
        "• Do not delete rows 1–4 (title, info, blank, headers).",
        "• Start entering data from row 5 onwards.",
        f"• Columns in orange are REQUIRED. Columns in blue are optional.",
        "• Choice columns have dropdown validation — select from the provided options.",
        "• For date columns, use DD/MM/YYYY format.",
        "",
        "BEFORE SUBMITTING:",
        "• Ensure all required fields are filled.",
        "• Do not merge cells or add extra sheets.",
        "• Save as .xlsx (not .xls or .csv).",
        "",
        f"Contact your PMU coordinator if you have questions.",
    ]
    for ri, line in enumerate(instructions, start=2):
        ws_inst.cell(row=ri, column=1, value=line)

    return _buf(wb)


# ── Validation Config (APP-002 compatible) ─────────────────────────────────────

def build_validation_config(job: MonitoringJob, config_name: str = None) -> bytes:
    """
    Export validation rules as a .pmuconfig JSON (importable into APP-002).
    Format matches config_manager.save_config() output structure.
    """
    name = config_name or f"{job.monitoring_name} Validation Rules"
    payload = {
        "app_id":  "APP-002",
        "name":    name,
        "slug":    name.lower().replace(" ", "_"),
        "version": 1,
        "saved":   "",
        "config":  {
            "workspace_slug":  job.workspace_slug,
            "project_code":    job.project_code,
            "source_filename": "",
            "artifact_id":     "",
            "transform_steps": [],
            "rules":           job.rule_defs,
            "column_mappings": [],
        },
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


# ── KPI Config (APP-003 compatible) ──────────────────────────────────────────

def build_kpi_config(job: MonitoringJob, config_name: str = None) -> bytes:
    """
    Export KPI definitions as a .pmuconfig JSON (importable into APP-003).
    """
    name = config_name or f"{job.monitoring_name} KPI Definitions"
    payload = {
        "app_id":  "APP-003",
        "name":    name,
        "slug":    name.lower().replace(" ", "_"),
        "version": 1,
        "saved":   "",
        "config":  {
            "workspace_slug":  job.workspace_slug,
            "project_code":    job.project_code,
            "source_filename": "",
            "artifact_id":     "",
            "agg_configs":     job.agg_defs,
            "kpi_defs":        job.kpi_defs,
            "rankings":        [],
            "variances":       [],
            "trends":          [],
        },
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


# ── Form Structure JSON ───────────────────────────────────────────────────────

def build_form_structure(job: MonitoringJob) -> bytes:
    """
    Export the form structure as JSON.
    Compatible with shared.add_form_items() for Google Forms creation.
    Also serves as a standalone reference document.
    """
    field_map = {f.field_id: f for f in job.fields}
    sections  = []

    for section in job.form_def.sections:
        if not section.enabled:
            continue
        questions = []
        for fid in section.field_ids:
            fld = field_map.get(fid)
            if fld and fld.enabled and not fld.is_calculated:
                q = {
                    "field_id":    fld.field_id,
                    "title":       fld.label or fld.name,
                    "description": fld.description,
                    "type":        fld.question_type(),
                    "required":    fld.required,
                    "choices":     fld.choices if fld.data_type in ("choice", "boolean") else [],
                }
                if fld.data_type == "boolean" and not q["choices"]:
                    q["choices"] = ["Yes", "No"]
                questions.append(q)
        sections.append({
            "title":       section.title,
            "description": section.description,
            "questions":   questions,
        })

    structure = {
        "title":       job.form_def.title or job.monitoring_name,
        "description": job.form_def.description,
        "project":     job.project_code,
        "entity_type": job.entity_type,
        "sections":    sections,
    }
    return json.dumps(structure, indent=2, ensure_ascii=False).encode("utf-8")


# ── Google Form Creation ──────────────────────────────────────────────────────

def create_google_form(job: MonitoringJob) -> dict:
    """
    Create a Google Form using the shared google_svc module.
    Returns form metadata dict or raises an exception if credentials unavailable.
    """
    import sys
    from pathlib import Path
    pmu_root = Path(__file__).parent.parent.parent.parent
    if str(pmu_root) not in sys.path:
        sys.path.insert(0, str(pmu_root))

    from shared import credentials_available, create_form, add_form_items

    if not credentials_available():
        raise RuntimeError("Google credentials not configured. "
                           "Place credentials.json in PMU_Tools/credentials/ to enable Google Forms.")

    form_meta = create_form(
        title       = job.form_def.title or job.monitoring_name,
        description = job.form_def.description or "",
    )

    field_map = {f.field_id: f for f in job.fields}
    sections  = []
    for section in job.form_def.sections:
        if not section.enabled:
            continue
        questions = []
        for fid in section.field_ids:
            fld = field_map.get(fid)
            if fld and fld.enabled and not fld.is_calculated:
                q = {
                    "title":       fld.label or fld.name,
                    "description": fld.description,
                    "type":        fld.question_type(),
                    "required":    fld.required,
                    "choices":     fld.choices if fld.data_type in ("choice", "boolean") else [],
                }
                if fld.data_type == "boolean" and not q["choices"]:
                    q["choices"] = ["Yes", "No"]
                questions.append(q)
        if questions:
            sections.append({
                "title":       section.title,
                "description": section.description,
                "questions":   questions,
            })

    if sections:
        add_form_items(form_meta["form_id"], sections)

    return form_meta
