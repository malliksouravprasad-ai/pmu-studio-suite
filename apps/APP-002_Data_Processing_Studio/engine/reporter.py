"""Output generator for APP-002 Data Processing Studio."""
import datetime
import io
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_APP_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PMU_ROOT = os.path.dirname(os.path.dirname(_APP_DIR))
for _p in [_PMU_ROOT, _APP_DIR]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from shared import register

_OUT_DIR = os.path.join(_APP_DIR, "outputs")
os.makedirs(_OUT_DIR, exist_ok=True)

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=13, bold=True, color="1F3864")
_INFO_FONT  = Font(italic=True, color="595959", size=9)
_LIGHT      = PatternFill("solid", fgColor="D9E1F2")
_GREEN      = PatternFill("solid", fgColor="C6EFCE")
_ORANGE     = PatternFill("solid", fgColor="FFEB9C")
_RED        = PatternFill("solid", fgColor="FFC7CE")
_EXCEL_MAX  = 1_048_572


def _hdr_cell(ws, value: str) -> WriteOnlyCell:
    c = WriteOnlyCell(ws, value=value)
    c.fill = _HDR_FILL
    c.font = _HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _set_col_widths(ws, columns):
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(50, max(10, len(str(col)) + 2))


def _write_df_stream(ws, df: pd.DataFrame):
    _set_col_widths(ws, df.columns)
    ws.append([_hdr_cell(ws, str(col)) for col in df.columns])
    for row_vals in df.head(_EXCEL_MAX).values.tolist():
        ws.append([None if (isinstance(v, float) and v != v) else v for v in row_vals])


def _buf(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _ts() -> str:
    return datetime.datetime.now().strftime("%d %b %Y %H:%M")


# ── Clean Dataset ─────────────────────────────────────────────────────────────

def save_clean_dataset(
    df: pd.DataFrame, project_code: str, artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    """Returns (file_path, BytesIO_buffer)."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Clean Dataset")

    title = WriteOnlyCell(ws, value=f"Clean Dataset — {artifact_id}")
    title.font = _TITLE_FONT
    ws.append([title])

    info = WriteOnlyCell(ws, value=f"Generated: {_ts()}  |  Rows: {len(df):,}  |  Columns: {len(df.columns)}")
    info.font = _INFO_FONT
    ws.append([info])
    ws.append([])

    _write_df_stream(ws, df)

    fname = f"{artifact_id}_clean_dataset.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue())
    buf.seek(0)
    return str(path), buf


# ── Transformation Log ────────────────────────────────────────────────────────

def save_transform_log(
    log_entries: list, project_code: str, artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transformation Log"

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"Transformation Log — {artifact_id}"
    ws["A1"].font  = _TITLE_FONT
    ws["A2"].value = f"Project: {project_code}  |  Generated: {_ts()}  |  Steps: {len(log_entries)}"
    ws["A2"].font  = _INFO_FONT

    cols = ["Step", "Operation", "Description", "Rows Before", "Rows After", "Rows Changed", "Cols Before", "Cols After"]
    for i, lbl in enumerate(cols, start=1):
        c = ws.cell(row=4, column=i, value=lbl)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 24

    for r, entry in enumerate(log_entries, start=5):
        ws.cell(row=r, column=1, value=entry.get("Step"))
        ws.cell(row=r, column=2, value=entry.get("Operation"))
        ws.cell(row=r, column=3, value=entry.get("Description"))
        ws.cell(row=r, column=4, value=entry.get("Rows Before"))
        ws.cell(row=r, column=5, value=entry.get("Rows After"))
        changed = entry.get("Rows Changed", 0)
        cc = ws.cell(row=r, column=6, value=changed)
        if changed != 0:
            cc.fill = _ORANGE
        ws.cell(row=r, column=7, value=entry.get("Cols Before"))
        ws.cell(row=r, column=8, value=entry.get("Cols After"))
        if r % 2 == 0:
            for col in [1, 2, 4, 5, 7, 8]:
                ws.cell(row=r, column=col).fill = _LIGHT

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    for col_letter in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 14

    fname = f"{artifact_id}_transformation_log.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue())
    buf.seek(0)
    return str(path), buf


# ── Validation Report ─────────────────────────────────────────────────────────

def save_validation_report(
    result_df: pd.DataFrame,
    val_results: list,
    all_exceptions: list,
    calc_log: list,
    project_code: str,
    artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    from .rule_executor import split_valid_invalid
    valid_df, invalid_df = split_valid_invalid(result_df)

    wb = Workbook(write_only=True)

    # Sheet 1: Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append([WriteOnlyCell(ws_sum, value=f"Validation Report — {artifact_id}")])
    ws_sum.append([WriteOnlyCell(ws_sum, value=f"Project: {project_code}  |  Generated: {_ts()}")])
    ws_sum.append([])
    ws_sum.append([
        _hdr_cell(ws_sum, "Rule"),
        _hdr_cell(ws_sum, "Type"),
        _hdr_cell(ws_sum, "Severity"),
        _hdr_cell(ws_sum, "Total Rows"),
        _hdr_cell(ws_sum, "Passed"),
        _hdr_cell(ws_sum, "Failed"),
        _hdr_cell(ws_sum, "Fail %"),
        _hdr_cell(ws_sum, "Status"),
    ])
    for rr in val_results:
        status_cell = WriteOnlyCell(ws_sum, value=rr.status)
        if rr.status == "PASS":
            status_cell.fill = _GREEN
        elif rr.status == "WARN":
            status_cell.fill = _ORANGE
        else:
            status_cell.fill = _RED
        ws_sum.append([
            WriteOnlyCell(ws_sum, value=rr.rule_name),
            WriteOnlyCell(ws_sum, value=rr.rule_type),
            WriteOnlyCell(ws_sum, value=rr.severity),
            WriteOnlyCell(ws_sum, value=rr.total),
            WriteOnlyCell(ws_sum, value=rr.passed),
            WriteOnlyCell(ws_sum, value=rr.failed),
            WriteOnlyCell(ws_sum, value=round(rr.fail_pct, 1)),
            status_cell,
        ])

    # Sheet 2: Exceptions
    ws_exc = wb.create_sheet("Exceptions")
    if all_exceptions:
        exc_df = pd.DataFrame(all_exceptions)
        _write_df_stream(ws_exc, exc_df)
    else:
        ws_exc.append([WriteOnlyCell(ws_exc, value="No exceptions found — all rows passed.")])

    # Sheet 3: Valid Data
    ws_valid = wb.create_sheet("Valid Data")
    if not valid_df.empty:
        _write_df_stream(ws_valid, valid_df)
    else:
        ws_valid.append([WriteOnlyCell(ws_valid, value="No valid rows.")])

    # Sheet 4: Invalid Data
    ws_inv = wb.create_sheet("Invalid Data")
    if not invalid_df.empty:
        _write_df_stream(ws_inv, invalid_df)
    else:
        ws_inv.append([WriteOnlyCell(ws_inv, value="No invalid rows.")])

    # Sheet 5: Calc Log
    if calc_log:
        ws_calc = wb.create_sheet("Calculated Columns")
        calc_df = pd.DataFrame(calc_log)
        _write_df_stream(ws_calc, calc_df)

    fname = f"{artifact_id}_validation_report.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue())
    buf.seek(0)
    return str(path), buf


# ── Registry ──────────────────────────────────────────────────────────────────

def register_outputs(
    artifact_id: str, project_code: str,
    clean_path: str, log_path: str = None, val_path: str = None,
    source_file: str = "", config_name: str = "", config_version: str = "",
) -> None:
    outputs = [(clean_path, "Clean Dataset")]
    if log_path:
        outputs.append((log_path, "Transformation Log"))
    if val_path:
        outputs.append((val_path, "Validation Report"))
    for path, rtype in outputs:
        register(
            artifact_id    = artifact_id,
            app_id         = "APP-002",
            project        = project_code,
            report_type    = rtype,
            output_file    = str(path),
            status         = "Generated",
            source_file    = source_file,
            config_name    = config_name,
            config_version = str(config_version),
        )


# ── Internal ──────────────────────────────────────────────────────────────────

def _out_path(fname: str, workspace_path: str = None):
    from pathlib import Path
    if workspace_path:
        base = Path(workspace_path) / "outputs"
        base.mkdir(parents=True, exist_ok=True)
    else:
        base = Path(_OUT_DIR)
    return base / fname
