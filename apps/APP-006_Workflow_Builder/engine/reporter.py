"""Output generator for APP-009 Workflow Builder."""
import dataclasses
import io
import json
import os
import sys
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_APP_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PMU_ROOT = os.path.dirname(os.path.dirname(_APP_DIR))
for _p in [_PMU_ROOT, _APP_DIR]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from shared import register
from .workflow_model import WorkflowJob, STATUS_EXCEL_FILLS
from .tracker_engine import get_tracking_matrix, compute_progress, get_pendency_list

_OUT_DIR = os.path.join(_PMU_ROOT, "outputs", "APP-009_Workflow_Builder")
os.makedirs(_OUT_DIR, exist_ok=True)

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=13, bold=True, color="1F3864")
_LIGHT      = PatternFill("solid", fgColor="D9E1F2")
_BOLD       = Font(bold=True)


def _hdr(ws, row, labels, start_col=1):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lbl)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w=12, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max_w, max(min_w, max_len + 2))


# ── Tracker Matrix sheet ──────────────────────────────────────────────────────

def _build_tracker_sheet(ws, job: WorkflowJob, matrix_df: pd.DataFrame):
    ws["A1"].value = f"Workflow Tracker — {job.workflow_name}"
    ws["A1"].font  = _TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"Entity: {job.entity_type}  |  "
            f"Stages: {len(job.stages)}  |  "
            f"Entities: {len(job.entities)}  |  "
            f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}").font = \
        Font(italic=True, color="595959")

    if matrix_df.empty:
        ws.cell(row=4, column=1, value="No entities or stages defined.")
        return

    # Header row
    _hdr(ws, 4, list(matrix_df.columns))
    ws.row_dimensions[4].height = 28

    stage_cols = list(matrix_df.columns)[1:]   # skip entity column

    for ri, (_, row) in enumerate(matrix_df.iterrows(), start=5):
        ws.cell(row=ri, column=1, value=row.iloc[0]).font = _BOLD
        for ci, stage_name in enumerate(stage_cols, start=2):
            status = row.get(stage_name, "Pending")
            cell   = ws.cell(row=ri, column=ci, value=status)
            fill_hex = STATUS_EXCEL_FILLS.get(status, "F2F2F2")
            cell.fill = PatternFill("solid", fgColor=fill_hex)
            cell.alignment = Alignment(horizontal="center")
            if status == "Overdue":
                cell.font = Font(bold=True, color="9C0006")
            elif status == "Completed":
                cell.font = Font(color="375623")

    ws.freeze_panes = "B5"
    _auto_width(ws)


# ── Stage Progress sheet ──────────────────────────────────────────────────────

def _build_progress_sheet(ws, progress_df: pd.DataFrame):
    ws["A1"].value = "Stage Progress Summary"
    ws["A1"].font  = _TITLE_FONT
    if progress_df.empty:
        ws.cell(row=3, column=1, value="No data.")
        return

    _hdr(ws, 3, list(progress_df.columns))
    ws.row_dimensions[3].height = 24

    for ri, (_, row) in enumerate(progress_df.iterrows(), start=4):
        pct = row.get("Completion %", 0)
        for ci, (col, val) in enumerate(row.items(), start=1):
            cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else None)
            if col == "Completion %":
                if pct >= 100: cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif pct >= 60: cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:           cell.fill = PatternFill("solid", fgColor="FFC7CE")
            elif ri % 2 == 0:
                cell.fill = _LIGHT
            if ci == 2:
                cell.font = _BOLD

    ws.freeze_panes = "A4"
    _auto_width(ws)


# ── Pendency Report sheet ─────────────────────────────────────────────────────

def _build_pendency_sheet(ws, pendency_df: pd.DataFrame, entity_type: str):
    ws["A1"].value = "Pendency Report"
    ws["A1"].font  = _TITLE_FONT

    if pendency_df.empty:
        ws.cell(row=3, column=1,
                value="All entities are on track — no pending items.").font = \
            Font(italic=True, color="375623")
        return

    _hdr(ws, 3, list(pendency_df.columns))
    ws.row_dimensions[3].height = 24

    for ri, (_, row) in enumerate(pendency_df.iterrows(), start=4):
        status = row.get("Status", "")
        fill_hex = STATUS_EXCEL_FILLS.get(status, "F2F2F2")
        sf = PatternFill("solid", fgColor=fill_hex)
        for ci, (col, val) in enumerate(row.items(), start=1):
            cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else "")
            cell.fill = sf
            if ci == 1:
                cell.font = _BOLD
            if status == "Overdue":
                cell.font = Font(bold=(ci == 1), color="9C0006")

    ws.freeze_panes = "A4"
    _auto_width(ws)


# ── Full Log sheet ────────────────────────────────────────────────────────────

def _build_log_sheet(ws, job: WorkflowJob):
    ws["A1"].value = "Full Tracking Log"
    ws["A1"].font  = _TITLE_FONT

    stage_map = {s.stage_id: s.name for s in job.stages}

    if not job.tracking:
        ws.cell(row=3, column=1, value="No tracking entries yet.")
        return

    _hdr(ws, 3, [job.entity_type, "Stage", "Status", "Updated On", "Remarks"])
    for ri, entry in enumerate(job.tracking, start=4):
        ws.cell(row=ri, column=1, value=entry.entity).font = _BOLD
        ws.cell(row=ri, column=2, value=stage_map.get(entry.stage_id, entry.stage_id))
        sf = STATUS_EXCEL_FILLS.get(entry.status, "F2F2F2")
        sc = ws.cell(row=ri, column=3, value=entry.status)
        sc.fill = PatternFill("solid", fgColor=sf)
        ws.cell(row=ri, column=4, value=entry.updated_on)
        ws.cell(row=ri, column=5, value=entry.remarks)
        if ri % 2 == 0:
            for ci in [1, 2, 4, 5]:
                ws.cell(row=ri, column=ci).fill = _LIGHT

    ws.freeze_panes = "A4"
    _auto_width(ws)


# ── Legend sheet ──────────────────────────────────────────────────────────────

def _build_legend(ws):
    ws["A1"].value = "Status Legend"
    ws["A1"].font  = _TITLE_FONT
    _hdr(ws, 3, ["Status", "Meaning"])
    items = [
        ("Completed",   "Entity has fully completed this stage."),
        ("In Progress", "Entity is actively working on this stage."),
        ("Pending",     "Stage not yet started."),
        ("Overdue",     "Stage is past its due date and not completed."),
        ("N/A",         "Stage does not apply to this entity."),
    ]
    for ri, (status, meaning) in enumerate(items, start=4):
        sf = STATUS_EXCEL_FILLS.get(status, "F2F2F2")
        ws.cell(row=ri, column=1, value=status).fill = PatternFill("solid", fgColor=sf)
        ws.cell(row=ri, column=2, value=meaning)
    _auto_width(ws)


# ── Main save functions ───────────────────────────────────────────────────────

def save_tracker(job: WorkflowJob, artifact_id: str) -> tuple:
    matrix_df   = get_tracking_matrix(job)
    progress_df = compute_progress(job)
    pendency_df = get_pendency_list(job)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_tracker  = wb.create_sheet("Tracker", 0)
    ws_progress = wb.create_sheet("Stage Progress")
    ws_pendency = wb.create_sheet("Pendency Report")
    ws_log      = wb.create_sheet("Full Log")
    ws_legend   = wb.create_sheet("Legend")

    _build_tracker_sheet(ws_tracker, job, matrix_df)
    _build_progress_sheet(ws_progress, progress_df)
    _build_pendency_sheet(ws_pendency, pendency_df, job.entity_type)
    _build_log_sheet(ws_log, job)
    _build_legend(ws_legend)

    fname = f"{artifact_id}_tracker.xlsx"
    path  = os.path.join(_OUT_DIR, fname)
    buf   = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return path, buf


def save_workflow_definition(job: WorkflowJob, artifact_id: str) -> tuple:
    data = {
        "artifact_id":   artifact_id,
        "workflow_name": job.workflow_name,
        "entity_type":   job.entity_type,
        "project_code":  job.project_code,
        "generated":     datetime.date.today().isoformat(),
        "stages":        [dataclasses.asdict(s) for s in job.stages],
        "entities":      job.entities,
        "tracking":      [dataclasses.asdict(t) for t in job.tracking],
    }
    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    fname = f"{artifact_id}_workflow_definition.json"
    path  = os.path.join(_OUT_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        f.write(json_str)

    buf = io.BytesIO(json_str.encode("utf-8"))
    return path, buf


def register_outputs(artifact_id, project_code, tracker_path, json_path):
    for path, rtype in [(tracker_path, "Tracker"),
                        (json_path, "Workflow Definition")]:
        register(
            artifact_id=artifact_id,
            app_id="APP-009",
            project=project_code,
            report_type=rtype,
            output_file=str(path),
            status="Generated",
        )
