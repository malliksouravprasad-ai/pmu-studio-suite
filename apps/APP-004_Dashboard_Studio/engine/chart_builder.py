"""Chart data preparation and openpyxl chart building for APP-007."""
from typing import List, Tuple
import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference

from .viz_model import ChartDef

_PANDAS_FUNC = {
    "sum": "sum", "avg": "mean", "count": "count",
    "min": "min", "max": "max", "distinct_count": "nunique",
}


def prepare_chart_data(df: pd.DataFrame, cdef: ChartDef) -> pd.DataFrame:
    """Aggregate and filter df for charting. Returns ready-to-plot DataFrame."""
    if not cdef.x_col or not cdef.y_cols:
        return pd.DataFrame()
    if cdef.x_col not in df.columns:
        return pd.DataFrame()

    # Pie chart: only one y-column
    y_cols = cdef.y_cols[:1] if cdef.chart_type == "pie" else cdef.y_cols
    present_y = [c for c in y_cols if c in df.columns]
    if not present_y:
        return pd.DataFrame()

    func = _PANDAS_FUNC.get(cdef.agg_func, "sum")
    result = (
        df[[cdef.x_col] + present_y]
        .groupby(cdef.x_col, as_index=False)
        .agg({c: func for c in present_y})
    )

    for c in present_y:
        result[c] = pd.to_numeric(result[c], errors="coerce").round(2)

    sort_col = cdef.sort_col if cdef.sort_col in result.columns else \
               (present_y[0] if present_y else "")
    if sort_col:
        result = result.sort_values(sort_col, ascending=cdef.sort_asc).reset_index(drop=True)

    if cdef.top_n > 0:
        result = result.head(cdef.top_n).reset_index(drop=True)

    return result


def write_chart_to_ws(
    ws_data,
    cdef: ChartDef,
    chart_df: pd.DataFrame,
    start_row: int,
) -> Tuple[object, int]:
    """Write chart data to ws_data and return (openpyxl_chart, next_available_row)."""
    if chart_df.empty:
        return None, start_row + 2

    y_cols = chart_df.columns.tolist()[1:]   # everything after x_col
    n_y = len(y_cols)

    # Write section header (row start_row)
    ws_data.cell(row=start_row, column=1, value=f"[Chart] {cdef.title}")

    # Write column headers (row start_row + 1)
    hdr_row = start_row + 1
    ws_data.cell(row=hdr_row, column=1, value=cdef.x_col)
    for ci, yc in enumerate(y_cols, start=2):
        ws_data.cell(row=hdr_row, column=ci, value=yc)

    # Write data rows
    data_start = hdr_row + 1
    for ri, (_, row) in enumerate(chart_df.iterrows(), start=data_start):
        ws_data.cell(row=ri, column=1, value=row.get(cdef.x_col))
        for ci, yc in enumerate(y_cols, start=2):
            val = row.get(yc)
            ws_data.cell(row=ri, column=ci, value=None if pd.isna(val) else val)

    data_end = data_start + len(chart_df) - 1
    next_row = data_end + 3  # gap between charts

    chart = _make_chart(ws_data, cdef, hdr_row, data_end, n_y)
    return chart, next_row


def _make_chart(ws_data, cdef: ChartDef, hdr_row: int, data_end: int, n_y: int):
    ct = cdef.chart_type.lower()

    if ct in ("column", "bar"):
        chart = BarChart()
        chart.type = "col" if ct == "column" else "bar"
        chart.grouping = "stacked" if cdef.stacked else "clustered"
    elif ct == "line":
        chart = LineChart()
    elif ct == "pie":
        chart = PieChart()
    elif ct == "area":
        chart = AreaChart()
        chart.grouping = "stacked" if cdef.stacked else "standard"
    else:
        chart = BarChart()
        chart.type = "col"

    chart.title = cdef.title
    chart.style = 10
    chart.width = 16   # cm
    chart.height = 10  # cm

    # Data reference (y-columns, headers included)
    data_ref = Reference(ws_data, min_col=2, max_col=n_y + 1,
                         min_row=hdr_row, max_row=data_end)
    chart.add_data(data_ref, titles_from_data=True)

    # Category reference (x-column, no header)
    cats_ref = Reference(ws_data, min_col=1,
                         min_row=hdr_row + 1, max_row=data_end)
    chart.set_categories(cats_ref)

    return chart
