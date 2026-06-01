"""Output generator for APP-007 Visualization Builder."""
import io
import os
import sys
import datetime
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_APP_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PMU_ROOT = os.path.dirname(os.path.dirname(_APP_DIR))
for _p in [_PMU_ROOT, _APP_DIR]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from shared import register
from .viz_model import DashboardJob, card_color
from .chart_builder import write_chart_to_ws

_OUT_DIR = os.path.join(_PMU_ROOT, "outputs", "APP-007_Visualization_Builder")
os.makedirs(_OUT_DIR, exist_ok=True)

# ── Styles ────────────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=14, bold=True, color="1F3864")
_LIGHT      = PatternFill("solid", fgColor="D9E1F2")
_BOLD       = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
_TOTAL_FONT = Font(bold=True, color="1F3864")

_STATUS_FILL = {
    "On Target": PatternFill("solid", fgColor="C6EFCE"),
    "Near":      PatternFill("solid", fgColor="FFEB9C"),
    "Critical":  PatternFill("solid", fgColor="FFC7CE"),
}


def _hdr(ws, row: int, labels: list, start_col: int = 1):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lbl)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max_w, max(min_w, max_len + 2))


def _write_df(ws, df: pd.DataFrame, start_row: int,
              is_total_col: str = "") -> int:
    """Write df to ws, return next row."""
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data.").font = \
            Font(italic=True, color="595959")
        return start_row + 2
    _hdr(ws, start_row, list(df.columns))
    ws.row_dimensions[start_row].height = 24
    for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        is_total = is_total_col and str(row.get(is_total_col, "")) == "TOTAL"
        for ci, (col, val) in enumerate(row.items(), start=1):
            cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else None)
            if is_total:
                cell.fill = _TOTAL_FILL
                cell.font = _TOTAL_FONT
            elif ri % 2 == 0:
                cell.fill = _LIGHT
            if ci == 1 and not is_total:
                cell.font = _BOLD
    return start_row + 1 + len(df)


# ── KPI card rendering ────────────────────────────────────────────────────────

_KPI_COLS = 4          # cards per row
_KPI_COL_WIDTH = 15    # column width in chars per card column (3 cols per card)
_KPI_TITLE_ROW_H = 22
_KPI_VALUE_ROW_H = 38
_KPI_COMP_ROW_H  = 20

_KPI_VALUE_FONT_SIZE = 20


def _draw_kpi_cards(ws, kpi_results: List[Dict[str, Any]], start_row: int) -> int:
    """Draw KPI cards on ws starting at start_row. Returns next available row."""
    if not kpi_results:
        return start_row

    cols_per_card = 3
    cards_per_row = _KPI_COLS

    for i, kpi in enumerate(kpi_results):
        row_in_group = i // cards_per_row
        col_in_group = i % cards_per_row

        r = start_row + row_in_group * 4   # 3 data rows + 1 gap
        c = col_in_group * cols_per_card + 1

        color = card_color(i, kpi.get("status", "neutral"))
        fill  = PatternFill("solid", fgColor=color)
        white = Font(color="FFFFFF", bold=True)
        white_sm = Font(color="FFFFFF", size=9)

        # Color all cells in the 3×3 card area
        for dr in range(3):
            for dc in range(cols_per_card):
                ws.cell(row=r + dr, column=c + dc).fill = fill

        # Row 1: Title
        ws.cell(row=r, column=c, value=kpi.get("title", "")).font = white_sm
        ws.row_dimensions[r].height = _KPI_TITLE_ROW_H

        # Row 2: Value
        val_cell = ws.cell(row=r + 1, column=c, value=kpi.get("formatted", "—"))
        val_cell.font = Font(color="FFFFFF", bold=True, size=_KPI_VALUE_FONT_SIZE)
        ws.row_dimensions[r + 1].height = _KPI_VALUE_ROW_H

        # Row 3: Comparison / target
        comp_str = ""
        if kpi.get("target") is not None:
            comp_str = f"Target: {kpi.get('target_fmt', '')}  {kpi.get('change_str', '')}"
        ws.cell(row=r + 2, column=c, value=comp_str).font = white_sm
        ws.row_dimensions[r + 2].height = _KPI_COMP_ROW_H

    # Set column widths for card columns
    n_card_cols = min(cards_per_row, len(kpi_results)) * cols_per_card
    for ci in range(1, n_card_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = _KPI_COL_WIDTH

    rows_used = ((len(kpi_results) - 1) // cards_per_row + 1) * 4
    return start_row + rows_used + 1   # +1 gap


# ── Dashboard sheet ───────────────────────────────────────────────────────────

def _build_dashboard(ws_dash, ws_chart_data, kpi_results, chart_data_map,
                     table_data_map, job: DashboardJob, artifact_id: str):
    # Title banner
    title_cell = ws_dash["A1"]
    title_cell.value = job.dashboard_title
    title_cell.font = Font(size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[1].height = 36

    ws_dash.cell(row=2, column=1,
                 value=f"Artifact: {artifact_id}  |  "
                 f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}").font = \
        Font(italic=True, color="595959")
    ws_dash.row_dimensions[2].height = 18

    current_row = 4

    # KPI cards
    if kpi_results:
        ws_dash.cell(row=current_row, column=1,
                     value="KEY PERFORMANCE INDICATORS").font = \
            Font(bold=True, size=11, color="1F3864")
        current_row += 1
        current_row = _draw_kpi_cards(ws_dash, kpi_results, current_row)

    # Charts
    if chart_data_map:
        ws_dash.cell(row=current_row, column=1,
                     value="CHARTS").font = Font(bold=True, size=11, color="1F3864")
        current_row += 1

        chart_row = current_row
        chart_data_row = 1   # row tracker in ws_chart_data

        for i, cdef in enumerate(job.charts):
            if not cdef.enabled:
                continue
            chart_df = chart_data_map.get(cdef.chart_id, pd.DataFrame())
            if chart_df.empty:
                continue

            chart_obj, chart_data_row = write_chart_to_ws(
                ws_chart_data, cdef, chart_df, chart_data_row)

            if chart_obj is None:
                continue

            # Alternate placement: 2 charts per row
            col_anchor = "A" if (i % 2 == 0) else "I"
            row_anchor = chart_row + (i // 2) * 22
            ws_dash.add_chart(chart_obj, f"{col_anchor}{row_anchor}")

        # Advance past all charts
        chart_rows = ((sum(1 for c in job.charts if c.enabled) + 1) // 2) * 22
        current_row += chart_rows

    # Tables (written in Dashboard sheet as compact formatted tables)
    if table_data_map:
        current_row += 1
        ws_dash.cell(row=current_row, column=1,
                     value="SUMMARY TABLES").font = Font(bold=True, size=11, color="1F3864")
        current_row += 1

        for tdef in job.tables:
            if not tdef.enabled:
                continue
            tdf = table_data_map.get(tdef.table_id, pd.DataFrame())
            if tdf.empty:
                continue

            ws_dash.cell(row=current_row, column=1,
                         value=tdef.title).font = Font(bold=True, color="1F3864", size=10)
            current_row += 1
            gc = tdef.group_cols[0] if tdef.group_cols else ""
            current_row = _write_df(ws_dash, tdf, current_row, is_total_col=gc)
            current_row += 2

    _auto_width(ws_dash)


# ── KPI Summary sheet ─────────────────────────────────────────────────────────

def _build_kpi_sheet(ws, kpi_results: List[Dict[str, Any]]):
    ws["A1"].value = "KPI Summary"
    ws["A1"].font = _TITLE_FONT
    if not kpi_results:
        ws.cell(row=3, column=1, value="No KPIs defined.")
        return

    _hdr(ws, 3, ["KPI", "Value", "Target", "Achievement %", "Change %", "Status"])
    for i, kpi in enumerate(kpi_results, start=4):
        val   = kpi.get("value")
        tgt   = kpi.get("target")
        achv  = round(val / tgt * 100, 1) if tgt and tgt != 0 and val is not None else None
        ws.cell(row=i, column=1, value=kpi.get("title")).font = _BOLD
        ws.cell(row=i, column=2, value=kpi.get("formatted"))
        ws.cell(row=i, column=3, value=kpi.get("target_fmt"))
        ws.cell(row=i, column=4, value=achv)
        ws.cell(row=i, column=5, value=kpi.get("change_str"))
        status = kpi.get("status", "neutral")
        status_label = {"good": "On Target", "warn": "Near Target",
                        "bad": "Critical", "neutral": "—"}.get(status, "—")
        sf = _STATUS_FILL.get(status_label)
        sc = ws.cell(row=i, column=6, value=status_label)
        if sf:
            sc.fill = sf
        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=i, column=c).fill = _LIGHT
    ws.freeze_panes = "A4"
    _auto_width(ws)


# ── Tables sheet ──────────────────────────────────────────────────────────────

def _build_tables_sheet(ws, table_data_map: Dict[str, pd.DataFrame], job: DashboardJob):
    ws["A1"].value = "Summary Tables"
    ws["A1"].font = _TITLE_FONT
    row = 3
    for tdef in job.tables:
        if not tdef.enabled:
            continue
        tdf = table_data_map.get(tdef.table_id, pd.DataFrame())
        ws.cell(row=row, column=1,
                value=tdef.title).font = Font(bold=True, color="1F3864", size=11)
        row += 1
        gc = tdef.group_cols[0] if tdef.group_cols else ""
        row = _write_df(ws, tdf, row, is_total_col=gc)
        row += 2
    _auto_width(ws)


# ── Main save functions ───────────────────────────────────────────────────────

def save_dashboard(
    kpi_results: List[Dict[str, Any]],
    chart_data_map: Dict[str, pd.DataFrame],
    table_data_map: Dict[str, pd.DataFrame],
    job: DashboardJob,
    artifact_id: str,
) -> tuple:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_dash       = wb.create_sheet("Dashboard", 0)
    ws_chart_data = wb.create_sheet("Chart Data")
    ws_kpi        = wb.create_sheet("KPI Summary")
    ws_tables     = wb.create_sheet("Summary Tables")

    _build_dashboard(ws_dash, ws_chart_data, kpi_results, chart_data_map,
                     table_data_map, job, artifact_id)
    _build_kpi_sheet(ws_kpi, kpi_results)
    _build_tables_sheet(ws_tables, table_data_map, job)

    fname = f"{artifact_id}_dashboard.xlsx"
    path  = os.path.join(_OUT_DIR, fname)
    buf   = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return path, buf


def save_dashboard_dataset(
    chart_data_map: Dict[str, pd.DataFrame],
    table_data_map: Dict[str, pd.DataFrame],
    job: DashboardJob,
    artifact_id: str,
) -> tuple:
    """Clean dataset workbook — one sheet per chart/table view, no styling."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for cdef in job.charts:
        if not cdef.enabled:
            continue
        df = chart_data_map.get(cdef.chart_id, pd.DataFrame())
        if df.empty:
            continue
        ws = wb.create_sheet(f"Chart_{cdef.title[:26]}"[:31])
        _write_df(ws, df, 1)
        _auto_width(ws)

    for tdef in job.tables:
        if not tdef.enabled:
            continue
        df = table_data_map.get(tdef.table_id, pd.DataFrame())
        if df.empty:
            continue
        ws = wb.create_sheet(f"Table_{tdef.title[:26]}"[:31])
        _write_df(ws, df, 1)
        _auto_width(ws)

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    fname = f"{artifact_id}_dashboard_dataset.xlsx"
    path  = os.path.join(_OUT_DIR, fname)
    buf   = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return path, buf


def register_outputs(artifact_id, project_code, dashboard_path, dataset_path):
    for path, rtype in [(dashboard_path, "Dashboard"),
                        (dataset_path, "Dashboard Dataset")]:
        register(
            artifact_id=artifact_id,
            app_id="APP-007",
            project=project_code,
            report_type=rtype,
            output_file=str(path),
            status="Generated",
        )
