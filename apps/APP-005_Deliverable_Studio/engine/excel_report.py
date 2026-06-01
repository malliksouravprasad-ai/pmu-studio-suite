"""Excel report generator for APP-008."""
import io
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .deliverable_model import ReportConfig, SectionDef

_NAVY   = "1F3864"
_WHITE  = "FFFFFF"
_LIGHT  = "D9E1F2"
_TOTAL  = "BDD7EE"

_HDR_FILL  = PatternFill("solid", fgColor=_NAVY)
_HDR_FONT  = Font(color=_WHITE, bold=True, size=10)
_TOTAL_FILL = PatternFill("solid", fgColor=_TOTAL)
_TOTAL_FONT = Font(bold=True, color=_NAVY)
_LIGHT_FILL = PatternFill("solid", fgColor=_LIGHT)
_BOLD       = Font(bold=True)
_THIN       = Side(style="thin", color="AAAAAA")
_BORDER     = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)


def _hdr(ws, row, labels, start_col=1):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lbl)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER


def _auto_width(ws, min_w=12, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max_w, max(min_w, max_len + 2))


def _write_df(ws, df: pd.DataFrame, start_row: int, is_total_col: str = "") -> int:
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data available.").font = \
            Font(italic=True, color="595959")
        return start_row + 2
    _hdr(ws, start_row, list(df.columns))
    ws.row_dimensions[start_row].height = 24
    for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        is_total = is_total_col and str(row.get(is_total_col, "")) == "TOTAL"
        for ci, (col, val) in enumerate(row.items(), start=1):
            cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else None)
            cell.border = _BORDER
            if is_total:
                cell.fill = _TOTAL_FILL; cell.font = _TOTAL_FONT
            elif ri % 2 == 0:
                cell.fill = _LIGHT_FILL
            if ci == 1 and not is_total:
                cell.font = _BOLD
    return start_row + 1 + len(df)


def _build_cover(ws, config: ReportConfig, artifact_id: str):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # Title banner
    title_cell = ws.cell(row=2, column=1, value=config.report_title)
    title_cell.font = Font(size=22, bold=True, color=_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 56

    # Colour adjacent cells
    for col in range(2, 6):
        c = ws.cell(row=2, column=col)
        c.fill = PatternFill("solid", fgColor=_NAVY)

    meta = [
        ("Programme", config.programme),
        ("Organisation", config.organization),
        ("Prepared by", config.author),
        ("Date", config.report_date),
        ("Artifact ID", artifact_id),
    ]
    for i, (label, value) in enumerate(meta, start=4):
        ws.cell(row=i, column=1, value=label).font = _BOLD
        ws.cell(row=i, column=2, value=value)
        ws.row_dimensions[i].height = 20

    if config.description:
        ws.cell(row=len(meta) + 5, column=1, value="Description").font = _BOLD
        ws.cell(row=len(meta) + 6, column=1, value=config.description).alignment = \
            Alignment(wrap_text=True)
        ws.row_dimensions[len(meta) + 6].height = 60


def save_excel_report(
    section_data: dict, config: ReportConfig, artifact_id: str
) -> tuple:
    import io
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_cover = wb.create_sheet("Cover", 0)
    _build_cover(ws_cover, config, artifact_id)

    for sdef in config.sections:
        if not sdef.enabled:
            continue
        ws = wb.create_sheet(sdef.title[:31])
        ws.cell(row=1, column=1, value=sdef.title).font = \
            Font(size=14, bold=True, color=_NAVY)
        ws.row_dimensions[1].height = 28

        if sdef.section_type == "narrative":
            ws.cell(row=3, column=1, value=sdef.narrative).alignment = \
                Alignment(wrap_text=True)
            ws.column_dimensions["A"].width = 80
            ws.row_dimensions[3].height = max(60, len(sdef.narrative) // 3)
        else:
            df = section_data.get(sdef.section_id, pd.DataFrame())
            gc = sdef.group_cols[0] if sdef.group_cols else ""
            _write_df(ws, df, 3, is_total_col=gc)
            ws.freeze_panes = "A4"
            _auto_width(ws)

    fname = f"{artifact_id}_report.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return fname, buf
