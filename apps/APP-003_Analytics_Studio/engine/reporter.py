"""Output generator for APP-003 Analytics Studio."""
import datetime
import io
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_APP_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PMU_ROOT = os.path.dirname(os.path.dirname(_APP_DIR))
for _p in [_PMU_ROOT, _APP_DIR]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from shared import register

_OUT_DIR    = os.path.join(_APP_DIR, "outputs")
os.makedirs(_OUT_DIR, exist_ok=True)

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=13, bold=True, color="1F3864")
_INFO_FONT  = Font(italic=True, color="595959", size=9)
_EXCEL_MAX  = 1_048_572


def _hdr_cell(ws, value):
    c = WriteOnlyCell(ws, value=str(value))
    c.fill = _HDR_FILL; c.font = _HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _write_df(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(40, max(10, len(str(col)) + 2))
    ws.append([_hdr_cell(ws, col) for col in df.columns])
    for row in df.head(_EXCEL_MAX).values.tolist():
        ws.append([None if (isinstance(v, float) and v != v) else v for v in row])


def _buf(wb):
    b = io.BytesIO(); wb.save(b); b.seek(0); return b


def _ts(): return datetime.datetime.now().strftime("%d %b %Y %H:%M")


def _out_path(fname: str, workspace_path: str = None):
    from pathlib import Path
    base = (Path(workspace_path) / "outputs") if workspace_path else Path(_OUT_DIR)
    base.mkdir(parents=True, exist_ok=True)
    return base / fname


# ── Aggregation Workbook ──────────────────────────────────────────────────────

def save_aggregation(
    agg_results: dict,  # {config_name: DataFrame}
    project_code: str, artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    wb = Workbook(write_only=True)
    for name, df in agg_results.items():
        sheet_name = name[:31] if len(name) > 31 else name
        ws = wb.create_sheet(sheet_name)
        _write_df(ws, df)

    fname = f"{artifact_id}_aggregation.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue()); buf.seek(0)
    return str(path), buf


# ── KPI Workbook ──────────────────────────────────────────────────────────────

def save_kpi_report(
    kpi_df: pd.DataFrame,
    kpi_summary: list,
    project_code: str, artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    wb = Workbook(write_only=True)

    # Sheet 1: KPI Summary
    ws_s = wb.create_sheet("KPI Summary")
    if kpi_summary:
        summary_df = pd.DataFrame(kpi_summary)
        _write_df(ws_s, summary_df)
    else:
        ws_s.append([WriteOnlyCell(ws_s, value="No KPIs calculated.")])

    # Sheet 2: KPI Dataset
    ws_d = wb.create_sheet("KPI Dataset")
    _write_df(ws_d, kpi_df)

    fname = f"{artifact_id}_kpi_report.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue()); buf.seek(0)
    return str(path), buf


# ── Analytics Workbook ────────────────────────────────────────────────────────

def save_analytics(
    rank_results: dict,
    var_results: dict,
    trend_results: dict,
    project_code: str, artifact_id: str,
    workspace_path: str = None,
) -> tuple:
    wb = Workbook(write_only=True)

    for name, df in rank_results.items():
        ws = wb.create_sheet(f"Rank-{name}"[:31])
        _write_df(ws, df)

    for name, df in var_results.items():
        ws = wb.create_sheet(f"Var-{name}"[:31])
        _write_df(ws, df)

    for name, df in trend_results.items():
        ws = wb.create_sheet(f"Trend-{name}"[:31])
        _write_df(ws, df)

    if not (rank_results or var_results or trend_results):
        ws = wb.create_sheet("No Results")
        ws.append([WriteOnlyCell(ws, value="No analytics results generated.")])

    fname = f"{artifact_id}_analytics.xlsx"
    path  = _out_path(fname, workspace_path)
    buf   = _buf(wb)
    path.write_bytes(buf.getvalue()); buf.seek(0)
    return str(path), buf


# ── Registry ──────────────────────────────────────────────────────────────────

def register_outputs(
    artifact_id: str, project_code: str,
    agg_path: str = None, kpi_path: str = None, analytics_path: str = None,
) -> None:
    for path, rtype in [
        (agg_path,       "Aggregation Dataset"),
        (kpi_path,       "KPI Report"),
        (analytics_path, "Analytics Report"),
    ]:
        if path:
            register(artifact_id=artifact_id, app_id="APP-003",
                     project=project_code, report_type=rtype,
                     output_file=str(path), status="Generated")
